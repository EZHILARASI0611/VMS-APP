"""
Export utilities for CSV and Excel reports.
"""
import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


VISITOR_EXPORT_HEADERS = [
    'Visitor Code', 'First Name', 'Last Name', 'Email', 'Phone', 'Company',
    'Purpose', 'Host Employee', 'Department', 'Status',
    'Check-In', 'Check-Out', 'Badge Number', 'Registered On'
]


def _format_row(visitor):
    """Format a visitor dict into export row."""
    return [
        visitor.get('visitor_code', ''),
        visitor.get('first_name', ''),
        visitor.get('last_name', ''),
        visitor.get('email', ''),
        visitor.get('phone', ''),
        visitor.get('company', ''),
        visitor.get('purpose', ''),
        visitor.get('host_name', ''),
        visitor.get('department_name', ''),
        visitor.get('status', ''),
        str(visitor.get('check_in_time') or ''),
        str(visitor.get('check_out_time') or ''),
        visitor.get('badge_number', ''),
        str(visitor.get('created_at') or ''),
    ]


def export_visitors_csv(visitors):
    """Return CSV string for visitor list."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(VISITOR_EXPORT_HEADERS)
    for v in visitors:
        writer.writerow(_format_row(v))
    return output.getvalue()


def export_visitors_excel(visitors):
    """Return Excel bytes for visitor list."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Visitors'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')

    for col, header in enumerate(VISITOR_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, visitor in enumerate(visitors, 2):
        for col_idx, value in enumerate(_format_row(visitor), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
